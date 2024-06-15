import pandas as pd
import openpyxl as xl 
import unicodedata
class Chaine:
    
    def __init__(self):
        self.nombreOccurence = {}
        
        
    def transformeChaineEnTableau(self, chaine):
        chaine = chaine.lower()
        chaine = chaine.split()
        chaine = unicodedata.normalize('NFKD', chaine).encode('ASCII', 'ignore')
        return chaine

    def occurences(self, chaine):
        for i in chaine:
            if (i in self.nombreOccurence):
                self.nombreOccurence[i] = self.nombreOccurence[i] + 1
            else:
                self.nombreOccurence[i] = 1
        return self.nombreOccurence
        
    def __str__(self):
        for i in self.nombreOccurence:
            print(i , ":" , self.nombreOccurence.get(i))
    
    def recupereDonnees(self, document):
        read = pd.read_excel(document)
        texte = read.to_string(index = False)
        return texte
    
    def createExcelFile(self) :
        workbook = xl.Workbook()
        sheet = workbook.active
        row = 1
        for key, _ in self.nombreOccurence.items() :
            sheet.cell(row=row, column=1, value=key)
            row += 1
            
        row = 1
        for _,value  in self.nombreOccurence.items():
            sheet.cell(row=row, column=2, value=value)
            row += 1
            
        workbook.save(filename='ouput.xlsx')
        return 0
        
chaine = Chaine()
tableau = chaine.recupereDonnees('/home/marion/Documents/ProgrammationC/Excel/workbook.xlsx')
chaineTransforme = chaine.transformeChaineEnTableau(tableau)
occurences = chaine.occurences(chaineTransforme)
print(occurences)
chaine.createExcelFile()
#chaine.__str__()
