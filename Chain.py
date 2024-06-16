import pandas as pd
import openpyxl as xl 
import unicodedata
class Chain:
    
    def __init__(self):
        self.occurrences = {}

    def board(self, chain):
        """Transforme une chaine de caractere en tableau de mots contenu dans la chaine de caractere

        Args:
            String chain : une chaine de caractere 

        Returns:
            String[] chain : un tableau de mots contenu dans la chaine de caractere sans majuscules et sans accents
        """
        chain = chain.lower()
        chain = unicodedata.normalize('NFKD', chain).encode('ASCII', 'ignore').decode('utf8')
        chain = chain.split()
        return chain
    
    def occurences(self, chain):
        """Identifie les occurrences 

        Args:
            String[] chain : un tableau de mots

        Returns:
            self.occurrences : un dictionnaire contenant les mots et leur frequence dans le tableau
        """
        for i in chain:
            if (i in self.occurrences):
                self.occurrences[i] = self.occurrences[i] + 1
            else:
                self.occurrences[i] = 1
        return self.occurrences
        
    def getDataInFile(self, document):
        """Recupere les mots contenu dans un document Excel

        Args:
            String document : chemin du fichier au sein des autres fichiers

        Returns:
            String text : une chaine de caractere contenu dans le fichier Excel
        """
        read = pd.read_excel(document)
        text = read.to_string(index = False)
        return text
    
    def createExcelFile(self) :
        """Creation d'un fichier Excel contenant les mots et les frequences du dictionnaire

        Returns:
            int : 0
        """
        workbook = xl.Workbook()
        active = workbook.active
        row = 1
        for key, _ in self.occurrences.items() :
            active.cell(row=row, column=1, value=key)
            row += 1
            
        row = 1
        for _,value  in self.occurrences.items():
            active.cell(row=row, column=2, value=value)
            row += 1
            
        workbook.save(filename='ouput.xlsx')
        return 0
        
chain = Chain()
tableau = chain.getDataInFile('/home/marion/Documents/ProgrammationC/Excel/workbook.xlsx')
chainTransforme = chain.board(tableau)
chain.occurences(chainTransforme)
chain.createExcelFile()
